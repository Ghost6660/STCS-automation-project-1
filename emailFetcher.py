import openpyxl
import win32com.client


SOURCE_FILE = "names.xlsx"
OUTPUT_FILE = "names_with_emails.xlsx"


def normalize_header(value: object) -> str:
    return str(value).replace("\ufeff", "").strip().strip('"').lower()


def find_header_column(ws, header_name: str) -> int:
    target = normalize_header(header_name)
    for cell in ws[1]:
        if normalize_header(cell.value) == target:
            return cell.column
    raise KeyError(f"Could not find a '{header_name}' column in sheet '{ws.title}'.")


def resolve_email(name: str, namespace) -> str:
    recipient = namespace.CreateRecipient(name)
    recipient.Resolve()

    if recipient.Resolved:
        try:
            exch_user = recipient.AddressEntry.GetExchangeUser()
            if exch_user:
                return exch_user.PrimarySmtpAddress
            return recipient.Address
        except Exception:
            return recipient.Address

    return "Not Found"


# Open Outlook
outlook = win32com.client.Dispatch("Outlook.Application")
namespace = outlook.GetNamespace("MAPI")

# Load Excel
wb = openpyxl.load_workbook(SOURCE_FILE)

for ws in wb.worksheets:
    owner_column = find_header_column(ws, "Owner")
    owner_email_column = find_header_column(ws, "Owner Email")

    for row in range(2, ws.max_row + 1):
        name = ws.cell(row=row, column=owner_column).value

        if not name:
            continue

        email = resolve_email(str(name), namespace)
        ws.cell(row=row, column=owner_email_column).value = email
        print(f"{name} -> {email}")

wb.save(OUTPUT_FILE)

print("Done!")